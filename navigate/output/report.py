# SPDX-FileCopyrightText: 2026 Fonden Mærsk Mc-Kinney Møller Center for Zero Carbon Shipping
# SPDX-License-Identifier: Apache-2.0

import csv
import logging
import os
from datetime import datetime
from enum import Enum

import numpy as np
import openpyxl as xl

from navigate.core import Node, assign_id
from navigate.core.enum_ import FileFormatID, ReportReduceID
from navigate.core.id_ import REPORT
from navigate.util import (
    attribute_to_setter,
    collapse_dict,
    collapse_tuple_dict,
    dates_to_days,
    is_single_dict,
    is_tuple_dict,
    retrieve_keys,
)

logger = logging.getLogger(__name__)

ROW_NODE = 1
ROW_ATTR = 2
ROW_KEY = 3
ROW_RESULT = 5


class NodeReport:
    def __init__(self):

        self.attributes = []
        self.getters = []
        self.reduce = []

    def add_property(self, attribute, reduce):
        if attribute not in self.attributes:
            self.attributes.append(attribute)
            self.getters.append(attribute_to_setter(attribute, method='get'))
            self.reduce.append(reduce)


class Report(Node):
    def __init__(self, name):
        super().__init__(name)

        self._type = REPORT

        # external properties
        self._directory = None  # str, either relative to deck directory or absolute path
        self._file_format = FileFormatID.XLSX  # Default format for backward compatibility

        # reports
        self._manager_reports = {}
        self._fleet_reports = {}
        self._levy_reports = {}
        self._plant_reports = {}
        self._port_reports = {}
        self._producer_reports = {}
        self._regulation_reports = {}
        self._vessel_reports = {}

        # static properties
        self._wb = None

    # external attributes set through the input deck -------------------------------------------------------------------
    def set_directory(self, directory):
        """
        Set the directory for where to export the report. Can be either a relative or absolute path.
        The directory will be created automatically if it doesn't exist.

        Examples
        --------
        - "./plots"

        Parameters
        ----------
        directory : str
            Relative or absolute path.
        """
        self._directory = directory

    def set_file_format(self, file_format):
        """
        Set the file format for report export.

        Parameters
        ----------
        file_format : str or FileFormatID
            Format for export ('XLSX' or 'CSV').
        """
        if isinstance(file_format, str):
            self._file_format = assign_id(file_format, FileFormatID)
        else:
            self._file_format = file_format

    # external commands called in the input deck -----------------------------------------------------------------------
    def add_property(self, attribute, reduce=None):
        self._add_property('global', self._manager_reports, attribute, reduce=reduce)

    def add_fleet_property(self, fleet_name, attribute, reduce=None):
        self._add_property(fleet_name, self._fleet_reports, attribute, reduce=reduce)

    def add_levy_property(self, levy_name, attribute, reduce=None):
        self._add_property(levy_name, self._levy_reports, attribute, reduce=reduce)

    def add_plant_property(self, plant_name, attribute, reduce=None):
        self._add_property(plant_name, self._plant_reports, attribute, reduce=reduce)

    def add_port_property(self, port_name, attribute, reduce=None):
        self._add_property(port_name, self._port_reports, attribute, reduce=reduce)

    def add_producer_property(self, producer_name, attribute, reduce=None):
        self._add_property(producer_name, self._producer_reports, attribute, reduce=reduce)

    def add_regulation_property(self, regulation_name, attribute, reduce=None):
        self._add_property(regulation_name, self._regulation_reports, attribute, reduce=reduce)

    def add_vessel_property(self, vessel_name, attribute, reduce=None):
        self._add_property(vessel_name, self._vessel_reports, attribute, reduce=reduce)

    # internal methods -------------------------------------------------------------------------------------------------
    def initialize(self):
        pass

    def start_export(self):
        if self._file_format == FileFormatID.XLSX:
            self._wb = xl.Workbook()
        else:  # CSV
            self._csv_data = {}

    def end_export(self, deck_directory, deck_name, dateline):
        if self._directory is not None:
            directory = os.path.join(deck_directory, self._directory)
        else:
            directory = deck_directory

        # Create directory if it doesn't exist
        os.makedirs(directory, exist_ok=True)

        if self._file_format == FileFormatID.XLSX:
            self._end_export_xlsx(directory, deck_name, dateline)
        else:
            self._end_export_csv(directory, deck_name, dateline)

    def _end_export_xlsx(self, directory, deck_name, dateline):
        """Export report in Excel format."""
        base_path = os.path.join(directory, '{}_{}.xlsx'.format(deck_name, self.name))

        # delete default sheet if data has been
        # written, otherwise keep it to avoid error
        if len(self._wb.sheetnames) > 1:
            del self._wb['Sheet']
            _export_date_time(self._wb, dateline)

        # save file with retry logic for locked files
        path = base_path
        max_attempts = 100

        for attempt in range(max_attempts):
            try:
                self._wb.save(path)
                if attempt > 0:
                    logger.warning(f"Saved report to alternative filename: {path}")
                break  # Success!
            except IOError:
                if attempt < max_attempts - 1:
                    # Generate alternative filename
                    path = _get_alternative_path(base_path, attempt + 1)
                else:
                    # Final attempt failed, re-raise the error
                    logger.error(f"Failed to save report after {max_attempts} attempts")
                    raise

    def _end_export_csv(self, directory, deck_name, dateline):
        """Export report in CSV format (one file per sheet)."""
        timeline = dates_to_days(dateline)

        for sheet_name, sheet_data in self._csv_data.items():
            base_path = os.path.join(directory, '{}_{}_{}.csv'.format(
                deck_name, self.name, sheet_name))

            path = base_path
            max_attempts = 100

            for attempt in range(max_attempts):
                try:
                    with open(path, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)

                        # Write headers
                        headers = ['Date', 'Time (days)'] + sheet_data['headers']
                        writer.writerow(headers)

                        # Write data rows
                        for i, date in enumerate(dateline):
                            row = [
                                str(date),  # Convert numpy datetime to string
                                timeline[i]
                            ]
                            # Extract value at index i from each column
                            row.extend([
                                col[i] if (hasattr(col, '__getitem__') and i < len(col))
                                else col
                                for col in sheet_data['columns']
                            ])
                            writer.writerow(row)

                    if attempt > 0:
                        logger.warning(f"Saved report to alternative filename: {path}")
                    break

                except IOError:
                    if attempt < max_attempts - 1:
                        path = _get_alternative_path(base_path, attempt + 1)
                    else:
                        logger.error(f"Failed to save report after {max_attempts} attempts")
                        raise

    def _export_properties(self, ws_or_name, nodes, extraction_dict):
        """

        Parameters
        ----------
        ws_or_name : Worksheet or str
            For Excel: Worksheet from openpyxl
            For CSV: Sheet name string
        nodes : dict[Node]
            Dict of all nodes of a certain type.
        extraction_dict : dict[str, NodeReport]
            Dict of all requested node reports.
        """

        if self._file_format == FileFormatID.XLSX:
            # Excel export - keep existing logic
            col = 3

            export = _prepare_export(nodes, extraction_dict)

            for node_name, (attributes, getters, reductions) in export.items():

                node = nodes[node_name]
                profile = node.profile

                properties = {}

                for attribute, getter, reduce in zip(attributes, getters, reductions):

                    try:
                        if hasattr(profile, getter):

                            property_ = getattr(profile, getter)()

                        else:

                            logger.error("Report '%s': Skipping property '%s' for node '%s': not a valid property.",
                                         self.name, attribute, node_name)
                            continue

                        if isinstance(property_, dict):

                            key1 = True if (reduce == ReportReduceID.FIRST or reduce == ReportReduceID.BOTH) else False
                            key2 = True if (reduce == ReportReduceID.SECOND or reduce == ReportReduceID.BOTH) else False

                            if property_:

                                if is_single_dict(property_):
                                    property_ = collapse_dict(property_, key=key1)

                                elif is_tuple_dict(property_):
                                    property_ = collapse_tuple_dict(property_, key1=key1, key2=key2)

                        elif isinstance(property_, list):

                            if reduce != ReportReduceID.NONE:
                                property_ = np.add.reduce(property_)

                        if property_ is not None:

                            properties[attribute] = property_

                    except Exception as e:
                        logger.error("Report '%s': Skipping property '%s' for node '%s': %s",
                                     self.name, attribute, node_name, e)
                        continue

                col = _export_node(ws_or_name, node, properties, col)
        else:
            # CSV export
            self._export_properties_csv(ws_or_name, nodes, extraction_dict)

    def _export_properties_csv(self, sheet_name, nodes, extraction_dict):
        """Export properties to CSV format."""

        export = _prepare_export(nodes, extraction_dict)

        headers = []
        columns = []

        for node_name, (attributes, getters, reductions) in export.items():

            node = nodes[node_name]
            profile = node.profile

            for attribute, getter, reduce in zip(attributes, getters, reductions):

                try:
                    if hasattr(profile, getter):

                        property_ = getattr(profile, getter)()

                    else:

                        logger.error("Report '%s': Skipping property '%s' for node '%s': not a valid property.",
                                     self.name, attribute, node_name)
                        continue

                    # Apply reduction logic (copied from _export_properties)
                    if isinstance(property_, dict):

                        key1 = True if (reduce == ReportReduceID.FIRST or reduce == ReportReduceID.BOTH) else False
                        key2 = True if (reduce == ReportReduceID.SECOND or reduce == ReportReduceID.BOTH) else False

                        if property_:

                            if is_single_dict(property_):
                                property_ = collapse_dict(property_, key=key1)

                            elif is_tuple_dict(property_):
                                property_ = collapse_tuple_dict(property_, key1=key1, key2=key2)

                    elif isinstance(property_, list):

                        if reduce != ReportReduceID.NONE:
                            property_ = np.add.reduce(property_)

                    if property_ is not None:
                        _flatten_to_csv(node_name, attribute, property_, headers, columns)

                except Exception as e:
                    logger.error("Report '%s': Skipping property '%s' for node '%s': %s",
                                 self.name, attribute, node_name, e)
                    continue

        # Store data for this sheet
        if headers:  # Only add if there's data
            self._csv_data[sheet_name] = {'headers': headers, 'columns': columns}

    def export_manager(self, manager):
        if self._manager_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Global')
                self._export_properties(ws, {manager.get_name(): manager}, self._manager_reports)
            else:
                self._export_properties('Global', {manager.get_name(): manager}, self._manager_reports)

    def export_fleets(self, fleets):
        if self._fleet_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Fleets')
                self._export_properties(ws, fleets, self._fleet_reports)
            else:
                self._export_properties('Fleets', fleets, self._fleet_reports)

    def export_levies(self, levies):
        if self._levy_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Levies')
                self._export_properties(ws, levies, self._levy_reports)
            else:
                self._export_properties('Levies', levies, self._levy_reports)

    def export_plants(self, plants):
        if self._plant_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Plants')
                self._export_properties(ws, plants, self._plant_reports)
            else:
                self._export_properties('Plants', plants, self._plant_reports)

    def export_ports(self, ports):
        if self._port_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Ports')
                self._export_properties(ws, ports, self._port_reports)
            else:
                self._export_properties('Ports', ports, self._port_reports)

    def export_producers(self, producers):
        if self._producer_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Producers')
                self._export_properties(ws, producers, self._producer_reports)
            else:
                self._export_properties('Producers', producers, self._producer_reports)

    def export_regulations(self, regulations):
        if self._regulation_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Regulations')
                self._export_properties(ws, regulations, self._regulation_reports)
            else:
                self._export_properties('Regulations', regulations, self._regulation_reports)

    def export_vessels(self, vessels):
        if self._vessel_reports:
            if self._file_format == FileFormatID.XLSX:
                ws = self._wb.create_sheet(title='Vessels')
                self._export_properties(ws, vessels, self._vessel_reports)
            else:
                self._export_properties('Vessels', vessels, self._vessel_reports)

    @staticmethod
    def _add_property(node_name, assignment_dict, attribute, reduce=None):

        internal_reduce = ReportReduceID.NONE

        if reduce is not None:
            internal_reduce = assign_id(reduce, ReportReduceID)

        if node_name not in assignment_dict:
            assignment_dict[node_name] = NodeReport()

        assignment_dict[node_name].add_property(attribute, internal_reduce)


def _get_alternative_path(base_path, counter):
    """
    Generate filename with counter suffix.

    Parameters
    ----------
    base_path : str
        Original file path (e.g., 'report.xlsx')
    counter : int
        Counter to append (e.g., 1 for 'report (1).xlsx')

    Returns
    -------
    str
        Alternative path with counter suffix

    Examples
    --------
    >>> _get_alternative_path('/path/to/file.xlsx', 1)
    '/path/to/file (1).xlsx'
    """
    base, ext = os.path.splitext(base_path)
    return f"{base} ({counter}){ext}"


def _prepare_export(nodes, extraction_dict):

    export = {}

    for key, report in extraction_dict.items():

        try:
            node_names = retrieve_keys(key, nodes)
        except KeyError:
            node_names = ()

        for node_name in node_names:

            if node_name not in export:
                export[node_name] = ([], [], [])

            export[node_name][0].extend(report.attributes)
            export[node_name][1].extend(report.getters)
            export[node_name][2].extend(report.reduce)

    return export


def _export_date_time(wb, dateline):

    timeline = dates_to_days(dateline)

    for ws in wb.worksheets:
        ws.cell(row=ROW_NODE, column=1).value = "Date"
        ws.cell(row=ROW_NODE, column=2).value = "Time (days)"
        _export_array(ws, '', dateline.astype(datetime), 1)
        _export_array(ws, '', timeline, 2)


def _export_node(ws, node, properties, col):

    first_col = col
    last_col = _export_properties(ws, properties, col) - 1

    # duplicate node name across each column header
    # if this is the first instance of the node export
    for col in range(first_col, last_col):
        ws.cell(row=ROW_NODE, column=col).value = node.get_name()

    return last_col


def _export_properties(ws, properties, col):

    for attribute, property_ in properties.items():

        try:
            if isinstance(property_, dict):
                col = _export_dict(ws, attribute, property_,  col)

            elif isinstance(property_, list):
                col = _export_list(ws, attribute, property_, col)

            elif isinstance(property_, np.ndarray):
                col = _export_array(ws, attribute, property_, col)

            else:
                logger.error("Skipping attribute '%s': unsupported type %s.", attribute, type(property_))
                continue
        except Exception as e:
            logger.error("Skipping attribute '%s': %s", attribute, e)
            continue

    return col + 1


def _export_dict(ws, attribute, property_, col, nested_dict=False):
    """

    Parameters
    ----------
    ws
    attribute
    property_
    col
    nested_dict : bool
        Used if a dict contains a dict

    Returns
    -------
    int
        The latest column index.
    """

    first_col = col
    offset = 1 if nested_dict else 0

    for key, value in property_.items():

        # for nested dicts
        initial_col = col

        if not isinstance(key, tuple):
            key = (key,)

        # TODO: DUPLICATE LIKE ROW ATTR
        # if not nested_dict:
        #     for k, key_ in enumerate(key):
        #         ws.cell(row=ROW_KEY + k + offset, column=col).value = _format_header(key_)

        if isinstance(value, dict):
            # in rare cases a dict may container another dict
            col = _export_dict(ws, *key, value, col, nested_dict=True)

            # duplicate first key across
            for c in range(initial_col, col):
                ws.cell(row=ROW_ATTR + 1, column=c).value = _format_header(key[0])

        else:
            col = _export_array(ws, attribute, value, col)

            for k, key_ in enumerate(key):
                ws.cell(row=ROW_KEY + k + offset, column=initial_col).value = _format_header(key_)

    # duplicate attribute name across each column header
    last_col = col

    if not nested_dict:
        for col in range(first_col, last_col):
            ws.cell(row=ROW_ATTR, column=col).value = attribute

    return last_col


def _export_list(ws, attribute, property_, col):

    first_col = col

    n = len(property_)
    indexes = range(1, n + 1)

    for idx, value in zip(indexes, property_):

        ws.cell(row=ROW_KEY, column=col).value = "Index {}".format(idx)
        col = _export_array(ws, attribute, value, col)

    # duplicate attribute name across each column header
    last_col = col

    for col in range(first_col, last_col):
        ws.cell(row=2, column=col).value = attribute

    return last_col


def _export_array(ws, attribute, property_, col):
    if attribute:
        ws.cell(row=ROW_ATTR, column=col).value = _format_header(attribute)

    return _export_time_series(ws, property_, col)


def _export_time_series(ws, series, col):
    n = series.size

    # write time-series
    for i in range(n):
        ws.cell(row=ROW_RESULT + i, column=col).value = series[i]

    return col + 1


def _format_header(header):
    if isinstance(header, Enum):
        return header.name
    else:
        return header


def _flatten_to_csv(node_name, attribute, property_, headers, columns):
    """
    Flatten a property structure into CSV headers and column data.

    Parameters
    ----------
    node_name : str
        Name of the node
    attribute : str
        Name of the attribute
    property_ : dict, list, or np.ndarray
        The property data to flatten
    headers : list
        List to append header names to (modified in-place)
    columns : list
        List to append column data to (modified in-place)
    """

    if isinstance(property_, dict):
        for key, value in property_.items():
            # Normalize key to tuple
            if not isinstance(key, tuple):
                key = (key,)

            key_str = '.'.join(str(_format_header(k)) for k in key)
            header = f"{node_name}.{attribute}.{key_str}"

            if isinstance(value, dict):
                # Nested dict - recursively flatten
                for sub_key, sub_value in value.items():
                    if not isinstance(sub_key, tuple):
                        sub_key = (sub_key,)
                    sub_key_str = '.'.join(str(_format_header(k)) for k in sub_key)
                    nested_header = f"{header}.{sub_key_str}"
                    headers.append(nested_header)
                    columns.append(sub_value)
            else:
                headers.append(header)
                columns.append(value)

    elif isinstance(property_, list):
        for idx, value in enumerate(property_, start=1):
            header = f"{node_name}.{attribute}.Index{idx}"
            headers.append(header)
            columns.append(value)

    elif isinstance(property_, np.ndarray):
        header = f"{node_name}.{attribute}"
        headers.append(header)
        columns.append(property_)

    else:
        logger.error("Skipping CSV attribute '%s' for node '%s': unsupported type %s.",
                     attribute, node_name, type(property_))
        return
