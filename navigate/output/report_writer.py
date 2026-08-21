# SPDX-FileCopyrightText: 2026 Fonden Mærsk Mc-Kinney Møller Center for Zero Carbon Shipping
# SPDX-License-Identifier: Apache-2.0

"""
Excel and CSV writing engine behind the Report node. The Report node collects which properties to
extract per node type; the functions here resolve those requests against the node profiles and
write the workbook or CSV files.
"""

import csv
import logging
import os
from datetime import datetime
from enum import Enum

import numpy as np
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet

from navigate.core.enum_ import ReportReduceID
from navigate.core.nodes.node import Node
from navigate.util import (
    attribute_to_setter,
    collapse_dict,
    collapse_tuple_dict,
    dates_to_days,
    is_single_dict,
    is_tuple_dict,
    retrieve_keys,
)

logger = logging.getLogger(__name__)

ROW_NODE = 1
ROW_ATTR = 2
ROW_KEY = 3
ROW_RESULT = 5


class NodeReport:
    def __init__(self) -> None:

        self.attributes: list[str] = []
        self.getters: list[str] = []
        self.reduce: list[ReportReduceID] = []

    def add_property(self, attribute: str, reduce: ReportReduceID) -> None:
        if attribute not in self.attributes:
            self.attributes.append(attribute)
            self.getters.append(attribute_to_setter(attribute, method='get'))
            self.reduce.append(reduce)


def write_xlsx_report(wb: xl.Workbook,
                      directory: str,
                      deck_name: str,
                      report_name: str,
                      dateline: np.ndarray
                      ) -> None:
    """
    Saves the workbook, retrying with alternative filenames while the target file is locked.

    Parameters
    ----------
    wb
        Workbook holding the exported sheets.
    directory
        Directory to save the file in.
    deck_name
        Name of the simulation deck, used in the filename.
    report_name
        Name of the Report node, used in the filename.
    dateline
        Dates of the simulation timeline.
    """

    base_path = os.path.join(directory, '{}_{}.xlsx'.format(deck_name, report_name))

    # delete default sheet if data has been
    # written, otherwise keep it to avoid error
    if len(wb.sheetnames) > 1:
        del wb['Sheet']
        _export_date_time(wb, dateline)

    # save file with retry logic for locked files
    path = base_path
    max_attempts = 100

    for attempt in range(max_attempts):
        try:
            wb.save(path)
            if attempt > 0:
                logger.warning(f"Saved report to alternative filename: {path}")
            break  # Success!
        except IOError:
            if attempt < max_attempts - 1:
                # Generate alternative filename
                path = _get_alternative_path(base_path, attempt + 1)
            else:
                # Final attempt failed, re-raise the error
                logger.error(f"Failed to save report after {max_attempts} attempts")
                raise


def write_csv_report(csv_data: dict,
                     directory: str,
                     deck_name: str,
                     report_name: str,
                     dateline: np.ndarray
                     ) -> None:
    """
    Writes one CSV file per sheet, retrying with alternative filenames while the target file is locked.

    Parameters
    ----------
    csv_data
        Per-sheet headers and columns collected by export_properties_csv.
    directory
        Directory to save the files in.
    deck_name
        Name of the simulation deck, used in the filenames.
    report_name
        Name of the Report node, used in the filenames.
    dateline
        Dates of the simulation timeline.
    """

    timeline = dates_to_days(dateline)

    for sheet_name, sheet_data in csv_data.items():
        base_path = os.path.join(directory, '{}_{}_{}.csv'.format(
            deck_name, report_name, sheet_name))

        path = base_path
        max_attempts = 100

        for attempt in range(max_attempts):
            try:
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)

                    # Write headers
                    headers = ['Date', 'Time (days)'] + sheet_data['headers']
                    writer.writerow(headers)

                    # Write data rows
                    for i, date in enumerate(dateline):
                        row = [
                            str(date),  # Convert numpy datetime to string
                            timeline[i]
                        ]
                        # Extract value at index i from each column
                        row.extend([
                            col[i] if (hasattr(col, '__getitem__') and i < len(col))
                            else col
                            for col in sheet_data['columns']
                        ])
                        writer.writerow(row)

                if attempt > 0:
                    logger.warning(f"Saved report to alternative filename: {path}")
                break

            except IOError:
                if attempt < max_attempts - 1:
                    path = _get_alternative_path(base_path, attempt + 1)
                else:
                    logger.error(f"Failed to save report after {max_attempts} attempts")
                    raise


def export_properties_xlsx(ws: Worksheet,
                           nodes: dict[str, Node],
                           extraction_dict: dict[str, NodeReport],
                           report_name: str
                           ) -> None:
    """
    Writes the requested properties of the given nodes into a worksheet.

    Parameters
    ----------
    ws
        Worksheet to write into.
    nodes
        Dict of all nodes of a certain type.
    extraction_dict
        Dict of all requested node reports.
    report_name
        Name of the Report node, used in log messages.
    """

    col = 3

    export = _prepare_export(nodes, extraction_dict)

    for node_name, (attributes, getters, reductions) in export.items():

        node = nodes[node_name]

        properties = dict(_extract_properties(node, attributes, getters, reductions, report_name))

        col = _export_node(ws, node, properties, col)


def export_properties_csv(sheet_name: str,
                          nodes: dict[str, Node],
                          extraction_dict: dict[str, NodeReport],
                          report_name: str,
                          csv_data: dict
                          ) -> None:
    """
    Flattens the requested properties of the given nodes into csv_data under the sheet name.

    Parameters
    ----------
    sheet_name
        Sheet name to store the flattened data under.
    nodes
        Dict of all nodes of a certain type.
    extraction_dict
        Dict of all requested node reports.
    report_name
        Name of the Report node, used in log messages.
    csv_data
        Per-sheet headers and columns, modified in-place.
    """

    export = _prepare_export(nodes, extraction_dict)

    headers = []
    columns = []

    for node_name, (attributes, getters, reductions) in export.items():

        node = nodes[node_name]

        for attribute, property_ in _extract_properties(node, attributes, getters, reductions, report_name):
            _flatten_to_csv(node_name, attribute, property_, headers, columns)

    # Store data for this sheet
    if headers:  # Only add if there's data
        csv_data[sheet_name] = {'headers': headers, 'columns': columns}


def _extract_properties(node: Node,
                        attributes: list[str],
                        getters: list[str],
                        reductions: list[ReportReduceID],
                        report_name: str):
    """
    Yields (attribute, property) pairs read from the node profile, with reductions applied.

    Parameters
    ----------
    node
        Node to read properties from.
    attributes
        Requested attribute names.
    getters
        Profile getter name per attribute.
    reductions
        Reduction to apply per attribute.
    report_name
        Name of the Report node, used in log messages.
    """

    profile = node.profile
    node_name = node.get_name()

    for attribute, getter, reduce in zip(attributes, getters, reductions):

        try:
            if hasattr(profile, getter):

                property_ = getattr(profile, getter)()

            else:

                logger.error("Report '%s': Skipping property '%s' for node '%s': not a valid property.",
                             report_name, attribute, node_name)
                continue

            if isinstance(property_, dict):

                key1 = True if (reduce == ReportReduceID.FIRST or reduce == ReportReduceID.BOTH) else False
                key2 = True if (reduce == ReportReduceID.SECOND or reduce == ReportReduceID.BOTH) else False

                if property_:

                    if is_single_dict(property_):
                        property_ = collapse_dict(property_, key=key1)

                    elif is_tuple_dict(property_):
                        property_ = collapse_tuple_dict(property_, key1=key1, key2=key2)

            elif isinstance(property_, list):

                if reduce != ReportReduceID.NONE:
                    property_ = np.add.reduce(property_)

            if property_ is not None:

                yield attribute, property_

        except Exception as e:
            logger.error("Report '%s': Skipping property '%s' for node '%s': %s",
                         report_name, attribute, node_name, e)
            continue


def _get_alternative_path(base_path: str, counter: int) -> str:
    """
    Generate filename with counter suffix.

    Parameters
    ----------
    base_path
        Original file path (e.g., 'report.xlsx')
    counter
        Counter to append (e.g., 1 for 'report (1).xlsx')

    Returns
    -------
    Alternative path with counter suffix, e.g. '/path/to/file (1).xlsx'.
    """
    base, ext = os.path.splitext(base_path)
    return f"{base} ({counter}){ext}"


def _prepare_export(nodes: dict[str, Node], extraction_dict: dict[str, NodeReport]) -> dict:

    export = {}

    for key, report in extraction_dict.items():

        try:
            node_names = retrieve_keys(key, nodes)
        except KeyError:
            node_names = ()

        for node_name in node_names:

            if node_name not in export:
                export[node_name] = ([], [], [])

            export[node_name][0].extend(report.attributes)
            export[node_name][1].extend(report.getters)
            export[node_name][2].extend(report.reduce)

    return export


def _export_date_time(wb: xl.Workbook, dateline: np.ndarray) -> None:

    timeline = dates_to_days(dateline)

    for ws in wb.worksheets:
        ws.cell(row=ROW_NODE, column=1).value = "Date"
        ws.cell(row=ROW_NODE, column=2).value = "Time (days)"
        _export_array(ws, '', dateline.astype(datetime), 1)
        _export_array(ws, '', timeline, 2)


def _export_node(ws: Worksheet, node: Node, properties: dict, col: int) -> int:

    first_col = col
    last_col = _write_properties(ws, properties, col) - 1

    # duplicate node name across each column header
    # if this is the first instance of the node export
    for col in range(first_col, last_col):
        ws.cell(row=ROW_NODE, column=col).value = node.get_name()

    return last_col


def _write_properties(ws: Worksheet, properties: dict, col: int) -> int:

    for attribute, property_ in properties.items():

        try:
            if isinstance(property_, dict):
                col = _export_dict(ws, attribute, property_,  col)

            elif isinstance(property_, list):
                col = _export_list(ws, attribute, property_, col)

            elif isinstance(property_, np.ndarray):
                col = _export_array(ws, attribute, property_, col)

            else:
                logger.error("Skipping attribute '%s': unsupported type %s.", attribute, type(property_))
                continue
        except Exception as e:
            logger.error("Skipping attribute '%s': %s", attribute, e)
            continue

    return col + 1


def _export_dict(ws: Worksheet, attribute: str, property_: dict, col: int, nested_dict: bool = False) -> int:
    """
    Writes a dict property into worksheet columns, recursing into nested dicts.

    Parameters
    ----------
    ws
        Worksheet to write into.
    attribute
        Name of the attribute, written as column header.
    property_
        Dict of values keyed by (tuples of) profile keys.
    col
        Column index to start writing at.
    nested_dict
        Used if a dict contains a dict

    Returns
    -------
    The latest column index.
    """

    first_col = col
    offset = 1 if nested_dict else 0

    for key, value in property_.items():

        # for nested dicts
        initial_col = col

        if not isinstance(key, tuple):
            key = (key,)

        # TODO: DUPLICATE LIKE ROW ATTR
        # if not nested_dict:
        #     for k, key_ in enumerate(key):
        #         ws.cell(row=ROW_KEY + k + offset, column=col).value = _format_header(key_)

        if isinstance(value, dict):
            # in rare cases a dict may container another dict
            col = _export_dict(ws, *key, value, col, nested_dict=True)

            # duplicate first key across
            for c in range(initial_col, col):
                ws.cell(row=ROW_ATTR + 1, column=c).value = _format_header(key[0])

        else:
            col = _export_array(ws, attribute, value, col)

            for k, key_ in enumerate(key):
                ws.cell(row=ROW_KEY + k + offset, column=initial_col).value = _format_header(key_)

    # duplicate attribute name across each column header
    last_col = col

    if not nested_dict:
        for col in range(first_col, last_col):
            ws.cell(row=ROW_ATTR, column=col).value = attribute

    return last_col


def _export_list(ws: Worksheet, attribute: str, property_: list, col: int) -> int:

    first_col = col

    n = len(property_)
    indexes = range(1, n + 1)

    for idx, value in zip(indexes, property_):

        ws.cell(row=ROW_KEY, column=col).value = "Index {}".format(idx)
        col = _export_array(ws, attribute, value, col)

    # duplicate attribute name across each column header
    last_col = col

    for col in range(first_col, last_col):
        ws.cell(row=2, column=col).value = attribute

    return last_col


def _export_array(ws: Worksheet, attribute: str, property_: np.ndarray, col: int) -> int:
    if attribute:
        ws.cell(row=ROW_ATTR, column=col).value = _format_header(attribute)

    return _export_time_series(ws, property_, col)


def _export_time_series(ws: Worksheet, series: np.ndarray, col: int) -> int:
    n = series.size

    # write time-series
    for i in range(n):
        ws.cell(row=ROW_RESULT + i, column=col).value = series[i]

    return col + 1


def _format_header(header):
    if isinstance(header, Enum):
        return header.name
    else:
        return header


def _flatten_to_csv(node_name: str, attribute: str, property_, headers: list, columns: list) -> None:
    """
    Flatten a property structure into CSV headers and column data.

    Parameters
    ----------
    node_name
        Name of the node
    attribute
        Name of the attribute
    property_
        The property data to flatten (dict, list, or np.ndarray).
    headers
        List to append header names to (modified in-place)
    columns
        List to append column data to (modified in-place)
    """

    if isinstance(property_, dict):
        for key, value in property_.items():
            # Normalize key to tuple
            if not isinstance(key, tuple):
                key = (key,)

            key_str = '.'.join(str(_format_header(k)) for k in key)
            header = f"{node_name}.{attribute}.{key_str}"

            if isinstance(value, dict):
                # Nested dict - recursively flatten
                for sub_key, sub_value in value.items():
                    if not isinstance(sub_key, tuple):
                        sub_key = (sub_key,)
                    sub_key_str = '.'.join(str(_format_header(k)) for k in sub_key)
                    nested_header = f"{header}.{sub_key_str}"
                    headers.append(nested_header)
                    columns.append(sub_value)
            else:
                headers.append(header)
                columns.append(value)

    elif isinstance(property_, list):
        for idx, value in enumerate(property_, start=1):
            header = f"{node_name}.{attribute}.Index{idx}"
            headers.append(header)
            columns.append(value)

    elif isinstance(property_, np.ndarray):
        header = f"{node_name}.{attribute}"
        headers.append(header)
        columns.append(property_)

    else:
        logger.error("Skipping CSV attribute '%s' for node '%s': unsupported type %s.",
                     attribute, node_name, type(property_))
        return
